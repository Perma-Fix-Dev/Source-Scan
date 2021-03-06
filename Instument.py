import tkinter as tk
from tkinter import filedialog
from tkinter import *
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os
import sys
import math

instruments = list()
SNs = list()
snRow = 1
avgRow = 26


class Instrument:

    def __init__(self, serial):
        self.SN = serial
        self.brickA = None
        self.brickB = None
        self.concreteA = None
        self.concreteB = None
        self.linoleumA = None
        self.linoleumB = None
        self.drywallA = None
        self.drywallB = None
        self.metalA = None
        self.metalB = None
        self.ceilingTileA = None
        self.ceilingTileB = None
        self.woodA = None
        self.woodB = None
        self.glassA = None
        self.glassB = None
        self.graniteA = None
        self.graniteB = None


def checklist(ws):
    col = 2

    while col < 1000:
        if len(instruments) == 0:
            inst = Instrument(ws.cell(column=col, row=snRow).value)
            instruments.append(inst)
            SNs.append(ws.cell(column=col, row=snRow).value)
        else:
            if ws.cell(column=col, row=snRow).value not in SNs:
                inst = Instrument(ws.cell(column=col, row=snRow).value)
                instruments.append(inst)
                SNs.append(ws.cell(column=col, row=snRow).value)

        col += 3


def setup(ws):
    ws.cell(column=1, row=1).value = "Serial Number"
    ws.cell(column=2, row=1).value = "Brick Alpha"
    ws.cell(column=3, row=1).value = "Brick Beta"
    ws.cell(column=4, row=1).value = "Concrete Alpha"
    ws.cell(column=5, row=1).value = "Concrete Beta"
    ws.cell(column=6, row=1).value = "Linoleum Alpha"
    ws.cell(column=7, row=1).value = "Linoleum Beta"
    ws.cell(column=8, row=1).value = "Drywall Alpha"
    ws.cell(column=9, row=1).value = "Drywall Beta"
    ws.cell(column=10, row=1).value = "Metal Alpha"
    ws.cell(column=11, row=1).value = "Metal Beta"
    ws.cell(column=12, row=1).value = "Ceiling Tile Alpha"
    ws.cell(column=13, row=1).value = "Ceiling Tile Beta"
    ws.cell(column=14, row=1).value = "Wood Alpha"
    ws.cell(column=15, row=1).value = "Wood Beta"
    ws.cell(column=16, row=1).value = "Glass Alpha"
    ws.cell(column=17, row=1).value = "Glass Beta"
    ws.cell(column=18, row=1).value = "Granite Alpha"
    ws.cell(column=19, row=1).value = "Granite Beta"


def auto_adjust(ws):
    for letter in range(1, ws.max_column+1):
        maximum_value = 0
        for cell in ws[get_column_letter(letter)]:
            val_to_check = len(str(cell.value))
            if val_to_check > maximum_value:
                maximum_value = val_to_check
        ws.column_dimensions[get_column_letter(letter)].width = maximum_value + 2


def autoCenter(ws):
    for col in ws.columns:
        for cell in col:
            # openpyxl styles aren't mutable,
            # so you have to create a copy of the style, modify the copy, then set it back
            cell.alignment = Alignment(horizontal='center')


def time_to_format(ws):
    auto_adjust(ws)
    autoCenter(ws)


def main(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    try:
        brickSheet = wb["Brick"]
        theBrickSheet = True
    except:
        theBrickSheet = False

    try:
        concreteSheet = wb["Concrete"]
        theConcreteSheet = True
    except:
        theConcreteSheet = False

    try:
        linoleumSheet = wb["Linoleum"]
        theLinoleumSheet = True
    except:
        theLinoleumSheet = False

    try:
        drywallSheet = wb["Drywall"]
        theDrywallSheet = True
    except:
        theDrywallSheet = False

    try:
        metalSheet = wb["Metal"]
        theMetalSheet = True
    except:
        theMetalSheet = False

    try:
        ceilingTileSheet = wb["CeilingTile"]
        theCeilingTileSheet = True
    except:
        theCeilingTileSheet = False

    try:
        woodSheet = wb["Wood"]
        theWoodSheet = True
    except:
        theWoodSheet = False

    try:
        glassSheet = wb["Glass"]
        theGlassSheet = True
    except:
        theGlassSheet = False

    try:
        graniteSheet = wb["Granite"]
        theGraniteSheet = True
    except:
        theGraniteSheet = False

    if theBrickSheet: checklist(brickSheet)
    if theConcreteSheet: checklist(concreteSheet)
    if theLinoleumSheet: checklist(linoleumSheet)
    if theDrywallSheet: checklist(drywallSheet)
    if theMetalSheet: checklist(metalSheet)
    if theCeilingTileSheet: checklist(ceilingTileSheet)
    if theWoodSheet: checklist(woodSheet)
    if theGlassSheet: checklist(glassSheet)
    if theGraniteSheet: checklist(graniteSheet)

    # Assign values
    if theBrickSheet:
        col = 2
        while col < 1000:
            for inst in instruments:
                if inst.SN == brickSheet.cell(column=col, row=snRow).value:
                    inst.brickA = brickSheet.cell(column=col, row=avgRow).value
                    inst.brickB = brickSheet.cell(column=col + 1, row=avgRow).value

                col += 3
                continue

    if theConcreteSheet:
        col = 2
        while col < 1000:
            for inst in instruments:
                if inst.SN == concreteSheet.cell(column=col, row=snRow).value:
                    inst.concreteA = concreteSheet.cell(column=col, row=avgRow).value
                    inst.concreteB = concreteSheet.cell(column=col + 1, row=avgRow).value

                col += 3
                continue
    if theLinoleumSheet:
        col = 2
        while col < 1000:
            for inst in instruments:
                if inst.SN == linoleumSheet.cell(column=col, row=snRow).value:
                    inst.linoleumA = linoleumSheet.cell(column=col, row=avgRow).value
                    inst.linoleumB = linoleumSheet.cell(column=col + 1, row=avgRow).value

                col += 3
                continue

    if theDrywallSheet:
        col = 2
        while col < 1000:
            for inst in instruments:
                if inst.SN == drywallSheet.cell(column=col, row=snRow).value:
                    inst.drywallA = drywallSheet.cell(column=col, row=avgRow).value
                    inst.drywallB = drywallSheet.cell(column=col + 1, row=avgRow).value

                col += 3
                continue

    if theMetalSheet:
        col = 2
        while col < 1000:
            for inst in instruments:
                if inst.SN == metalSheet.cell(column=col, row=snRow).value:
                    inst.metalA = metalSheet.cell(column=col, row=avgRow).value
                    inst.metalB = metalSheet.cell(column=col + 1, row=avgRow).value

                col += 3
                continue

    if theCeilingTileSheet:
        col = 2
        while col < 1000:
            for inst in instruments:
                if inst.SN == ceilingTileSheet.cell(column=col, row=snRow).value:
                    inst.ceilingTileA = ceilingTileSheet.cell(column=col, row=avgRow).value
                    inst.ceilingTileB = ceilingTileSheet.cell(column=col + 1, row=avgRow).value

                col += 3
                continue

    if theWoodSheet:
        col = 2
        while col < 1000:
            for inst in instruments:
                if inst.SN == woodSheet.cell(column=col, row=snRow).value:
                    inst.woodA = woodSheet.cell(column=col, row=avgRow).value
                    inst.woodB = woodSheet.cell(column=col + 1, row=avgRow).value

                col += 3
                continue

    if theGlassSheet:
        col = 2
        while col < 1000:
            for inst in instruments:
                if inst.SN == glassSheet.cell(column=col, row=snRow).value:
                    inst.glassA = glassSheet.cell(column=col, row=avgRow).value
                    inst.glassB = glassSheet.cell(column=col + 1, row=avgRow).value

                col += 3
                continue

    if theGraniteSheet:
        col = 2
        while col < 1000:
            for inst in instruments:
                if inst.SN == graniteSheet.cell(column=col, row=snRow).value:
                    inst.graniteA = graniteSheet.cell(column=col, row=avgRow).value
                    inst.graniteB = graniteSheet.cell(column=col + 1, row=avgRow).value

                col += 3
                continue

    # Print to final sheet
    row = 2
    finalSheet = wb.create_sheet('Averages by Material')
    setup(finalSheet)

    for inst in instruments:
        try:
            if inst.SN is not None:
                finalSheet.cell(column=1, row=row).value = inst.SN
                if theBrickSheet: finalSheet.cell(column=2, row=row).value = round(inst.brickA + 0.000000000000000000001, 0)
                if theBrickSheet: finalSheet.cell(column=3, row=row).value = round(inst.brickB + 0.000000000000000000001, 0)
                if theConcreteSheet: finalSheet.cell(column=4, row=row).value = round(inst.concreteA + 0.000000000000000000001, 0)
                if theConcreteSheet: finalSheet.cell(column=5, row=row).value = round(inst.concreteB + 0.000000000000000000001, 0)
                if theLinoleumSheet: finalSheet.cell(column=6, row=row).value = round(inst.linoleumA + 0.000000000000000000001, 0)
                if theLinoleumSheet: finalSheet.cell(column=7, row=row).value = round(inst.linoleumB + 0.000000000000000000001, 0)
                if theDrywallSheet:  finalSheet.cell(column=8, row=row).value = round(inst.drywallA + 0.000000000000000000001, 0)
                if theDrywallSheet: finalSheet.cell(column=9, row=row).value = round(inst.drywallB + 0.000000000000000000001, 0)
                if theMetalSheet: finalSheet.cell(column=10, row=row).value = round(inst.metalA + 0.000000000000000000001, 0)
                if theMetalSheet: finalSheet.cell(column=11, row=row).value = round(inst.metalB + 0.000000000000000000001, 0)
                if theCeilingTileSheet: finalSheet.cell(column=12, row=row).value = round(inst.ceilingTileA + 0.000000000000000000001, 0)
                if theCeilingTileSheet: finalSheet.cell(column=13, row=row).value = round(inst.ceilingTileB + 0.000000000000000000001, 0)
                if theWoodSheet: finalSheet.cell(column=14, row=row).value = round(inst.woodA + 0.000000000000000000001, 0)
                if theWoodSheet: finalSheet.cell(column=15, row=row).value = round(inst.woodB + 0.000000000000000000001, 0)
                if theGlassSheet: finalSheet.cell(column=16, row=row).value = round(inst.glassA + 0.000000000000000000001, 0)
                if theGlassSheet: finalSheet.cell(column=17, row=row).value = round(inst.glassB + 0.000000000000000000001, 0)
                if theGraniteSheet: finalSheet.cell(column=18, row=row).value = round(inst.graniteA + 0.000000000000000000001, 0)
                if theGraniteSheet: finalSheet.cell(column=19, row=row).value = round(inst.graniteB + 0.000000000000000000001, 0)

        except:
            print("Error: with inputting inst " + str(inst.SN))

        row += 1

    time_to_format(finalSheet)
    wb.save(filepath)


def select_path():
    global path
    curr_directory = os.getcwd()
    filename = filedialog.askopenfilename(initialdir=curr_directory, title="Select File")
    path.set(filename)


def resource_path(relative_path):
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def dummy_main():
    global path
    main(path.get())


# GUI
root = Tk()
root.resizable(width=False, height=False)
root.title('Data Analysis')
root.geometry('220x100')
image = PhotoImage(file=resource_path("images.png"))
path = StringVar()

label = tk.Label(root, text="File Path:")
label.place(x=0, y=5)
entry = tk.Entry(root, width=20, text=path)
entry.place(x=67, y=7)
button1 = tk.Button(root, image=image, width=20, height=20, command=select_path)
button1.place(x=190, y=3)
button1 = tk.Button(root, text="Go", command=dummy_main)
button1.place(x=100, y=30)

root.mainloop()
